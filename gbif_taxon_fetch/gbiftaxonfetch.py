import urllib.parse, urllib.request, json, sys, time, pickle, atexit
from pathlib import Path
# Non-std library modules
import toml, openpyxl 
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
paleblue = Color(indexed=44)

class kotka_excel:
    """First row, first col = 1"""
    def __init__(self,filename):     
        self.fp = Path(filename)
        self.wb = openpyxl.load_workbook(self.fp)
        self.sheet = self.wb.active
        self.name2col = {}
        self._update_name2column()
        self.crow = 3 # Current row: skip the first two header lines
    def _getcell(self,colno,rowno): # Get value
        return self.sheet.cell(column=colno,row=rowno).value 
    def _getncell(self,colname,rowno): # Get value by column name
        colno = self.name2col[colname]
        return self._getcell(colno,rowno)
    def _setcell(self,colno,rowno,value): # Get value
        self.sheet.cell(column=colno,row=rowno).value = value
    def _setncell(self,colname,rowno,value): # Get value by column name
        colno = self.name2col[colname]
        self._setcell(colno,rowno,value)
    def save(self): self.wb.save(str(self.fp))
    def close(self): pass
    def hascolumn(self,colname): return colname in self.name2col
    def addcolumn(self,position,title="New column"):
        """Add column. Note: First column is 1 etc."""
        if title in self.name2col: 
            raise ValueError(f"Column name {title} does already exist")
        self.sheet.insert_cols(position) 
        self._setcell(position,1,title)
        self._update_name2column()        
    def _update_name2column(self):
        _cotitles = tuple(( x.value for x in self.sheet[1]) )
        _ind = tuple(range(1,len(_cotitles)+1))
        self.name2col = { c:i for (c,i) in zip(_cotitles,_ind) }
    # Support simple iteration over lines
    def __next__(self): self.crow += 1
    def next(self): self.__next__()
    def iterget(self,colno): return self._getcell(colno,self.crow)
    def iterset(self,colno,value): self._setncell(colno,self.crow,value)
    def iternget(self,colname): return self._getncell(colname,self.crow)
    def iternset(self,colname,value): self._setncell(colname,self.crow,value)
    def iternfill(self,colname,color): 
        colno = self.name2col[colname]    
        cell = self.sheet.cell(column=colno,row=self.crow)        
        cell.fill = PatternFill("solid", fgColor=color)
    def end(self): 
        if self.crow > self.sheet.max_row: return True
        else: return False
    def newpath(self,newfp):
        self.fp = Path(newfp)    
    def add_note_to_name(self,note,sep="_"):
        newfp = self.fp.with_name(self.fp.stem + sep + note + self.fp.suffix)        
        self.newpath(newfp)
        

class gbif_cache:
    """Simple locally caching data source for GBIF name data. Stores data in a Python binary pickle."""
    _return_on_failure = {}
    def __init__(self,filename):
        self.data = {}
        self.fn = Path(filename)
    def open(self):
        # Read pickled data if exists
        print("Loading GBIF cache data from", self.fn)
        try:
            with open(self.fn,"rb") as f: self.data = pickle.load(f)
        except FileNotFoundError:
            print("No pickle file found")
            self.data = {}
    def flush(self):
        self.data = {}    
    def close(self):        
        with open(self.fn,"wb") as f: pickle.dump(self.data,f)
    def getGBIFData(self,taxonname,classname):# classname = limit name searches to a taxon
        "returns (name,id,family)"
        if not taxonname: return self._return_on_failure
        gbifURLbase = "https://api.gbif.org/v1/species/match?class=%s&name=%s&strict=True"
        gbifURLvalidname = "https://api.gbif.org/v1/species/%i"
        searchname = urllib.parse.quote(taxonname)
        print("Querying GBIF for %s" % searchname)
        if searchname in self.data: # Check if name is in cached GBIF data
            print("Found in cache")
            return self.data[searchname]
        else: # If not in cache, then call GBIF API to check name
            self.data[searchname] = self._return_on_failure 
            time.sleep(1.0) # slow down not to overload the Web query
            GBIFurl = gbifURLbase % (classname,searchname)
            with urllib.request.urlopen(GBIFurl) as f:
                res = json.loads(str(f.read(),"utf-8"))
                if res['matchType'] == 'NONE': return self.data[searchname] # No match
                else: 
                    # this setion would replace with valid name 
                    # if (res["synonym"] is True): # lookup recommended name for a synonym
                            # validkey = res["speciesKey"]
                            # with urllib.request.urlopen(gbifURLvalidname % validkey) as f2:
                                # res = json.loads(str(f2.read(),"utf-8"))
                    self.data[searchname] = {\
                    "key": res.get("usageKey",""),
                    "name": res["scientificName"],
                    "cname": res["canonicalName"],
                    "family": res.get("family",""),
                    "rank": res.get("rank","").lower(),
                    }
            return self.data[searchname]

def splitGBIFname(n,cn):
    # Simple implementation, may fail: takes name and canonical name
    # Return name, author
    if not n.strip(): return (n,"") 
    else: return ( cn.strip() , n.replace(cn,"").strip() )
        
def ReadConfig(configfile): 
    # Todo, check error states if opening fails (malformatted file & other errors)
    # Check that working directory etc. exist?
    configfile = Path(configfile)
    if not configfile.exists(): return False
    config = toml.load(configfile)
    return config

if __name__ == '__main__':
    try:
        # TODO: THis should not be hardcoded
        install_dir = Path("P:/h978/insect/Kahanpää/SMALL_TOOLS/gbif_taxon_fetch")
        configfile = install_dir / Path("gbiftaxonfetch.toml") 
        gbifsource = kfile = None
        print(f"Reading config file {configfile}")
        conf = ReadConfig(configfile)
        col_species = conf["excel"]["col_taxon_name"]
        col_author = conf["excel"]["col_taxon_author"]
        col_rank = conf["excel"]["col_taxon_rank"]
        col_family = conf["excel"]["family_column_title"]
        col_gbifname = conf["excel"]["gbifname_column_title"]
        if ( len(sys.argv) <= 1 ): 
            print("No input data file, exiting...")
            sys.exit()
        # Open (and if requested, flush) GBIF cache file
        if conf["gbif"]["flush_GBIF_cache"]:
            print("Flushing GBIF cache ... ",end='', flush=True)
            try:
                gbifsource = gbif_cache(conf["gbif"]["cachefile"])
                gbifsource.open()
                gbifsource.flush()
                print("clear")
            finally:
                gbifsource.close()            
        gbifsource = gbif_cache( conf["gbif"]["cachefile"] )
        gbifsource.open()
        # Loop over input files
        for inputfn in sys.argv[1:]:
            kfile = kotka_excel(inputfn)
            col_add_pos = conf["excel"].get("add_gbifname_column_at", None)
            if col_add_pos is not None:
                kfile.addcolumn(col_add_pos,col_gbifname)
            col_add_pos = conf["excel"].get("add_family_column_at", None)
            if col_add_pos is not None:
                kfile.addcolumn(col_add_pos,col_family)
            while not kfile.end():
                author = kfile.iternget(col_author)
                if author: # Skip if has author
                    kfile.next(); continue               
                name = kfile.iternget(col_species)
                result = gbifsource.getGBIFData( name,conf["taxonomy"]["class"] )
                # Result = (fullname, GBIF_id, family)
                if result: #Has some output
                    gbifname,author = splitGBIFname( result["name"],result["cname"] )
                    if kfile.hascolumn(col_author):
                        kfile.iternset(col_author,author)
                        kfile.iternfill(col_author,paleblue)
                    if kfile.hascolumn(col_rank):
                        kfile.iternset(col_rank,result["rank"])
                        kfile.iternfill(col_rank,paleblue)                    
                    if kfile.hascolumn(col_gbifname):
                        kfile.iternset(col_gbifname,gbifname)
                        kfile.iternfill(col_gbifname,paleblue)
                    if kfile.hascolumn(col_family):
                        kfile.iternset(col_family,result["family"])
                        kfile.iternfill(col_family,paleblue)
                kfile.next()
            done_marker = "gbif"
            sep = "_"
            kfile.add_note_to_name(done_marker,sep)
            try: kfile.save()
            except PermissionError: 
                print(f"WARNING: Could not write into file {kfile.fp}")                
    except Exception as err: # Catch any Exception
        raise err

@atexit.register
def goodbye(): # Close open files
        if gbifsource: gbifsource.close()
        if kfile: kfile.close()
# TODO: 
# Write README
# CALL LAJITIETOKESKUS API (need permission)
# CALL GBIF API for laji.fi failed cases 